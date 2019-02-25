from tkinter import *
from openpyxl import Workbook

handler1 = Workbook()
sheet1 = handler1.active
i=2
def abc():
  global i 
  sheet1.cell(row=1, column=1).value = "Name"
  sheet1.cell(row=i, column=1).value = text_entry1.get()

  sheet1.cell(row=1, column=2).value = "Class"
  sheet1.cell(row=i, column=2).value = text_entry2.get()
 
  sheet1.cell(row=1, column=3).value = "Rollno"
  sheet1.cell(row=i, column=3).value = text_entry3.get()
 
  i= i+1 
  handler1.save("test.xlsx")



#----------------main---------------
window = Tk()  # important line , it create a window
window.title("Student Result") # window title

#--------------LABLE 1 ----------
lb = Label(window,text="- - -Student Form- - -",bg="yellow",fg="black",font="none 12 bold")
lb.grid(row=1,column=0,sticky=S)

#-------------LABLE 2 -NAME OF STUDENT -----------
lb = Label(window,text="NAME",fg="black",font="none 10 bold")
lb.grid(row=2,column=0,sticky=S)
#------create a text entry box-----
text_entry1 = Entry(window,width=20,bg="white")
text_entry1.grid(row=3,column=0,sticky=S)


#-------------LABLE 3---CLASS OF STUDENT-----------
lb = Label(window,text="CLASS",fg="black",font="none 10 bold")
lb.grid(row=4,column=0,sticky=S)
#------create a text entry box-----
text_entry2 = Entry(window,width=20,bg="white")
text_entry2.grid(row=5,column=0,sticky=S)
#----------------------------------------------------


#-------------LABLE 4---ROLL NUMBER-----------
lb = Label(window,text="ROLL-NO",fg="black",font="none 10 bold")
lb.grid(row=6,column=0,sticky=S)
#------create a text entry box-----
text_entry3 = Entry(window,width=20,bg="white")
text_entry3.grid(row=7,column=0,sticky=S)
#----------------------------------------------------


# add a submit button
# click is a  user defined function
b= Button(window,text="SUBMIT",width=6,command=abc)
b.grid(row=10,column=0,sticky=S)

window.mainloop()
