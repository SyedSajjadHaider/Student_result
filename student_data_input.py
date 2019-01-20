from openpyxl import Workbook
handler1 = Workbook()
sheet1 = handler1.active

#----------------------------------------------
sheet1.cell(row=1, column=1).value = "Name"
sheet1.cell(row=1, column=2).value = "Roll No"
sheet1.cell(row=1, column=3).value = "Marks"

#----------------------------------------------
var = 1
i=2
while var == 1:
    
    # learning typecasting is necessary as by default
    # it will take as string not integer
    sheet1.cell(row=i, column=1).value = input("Enter Name \n")
    sheet1.cell(row=i, column=2).value = input("Enter rollno \n")
    sheet1.cell(row=i, column=3).value = int(input("Enter Marks\n"))
    var = int(input("continue: press 1 else 0\n"))
    i = i+1


#-----------------------------------------------

handler1.save("test.xlsx")
