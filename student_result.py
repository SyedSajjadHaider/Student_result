# read sheet
# customize sheet
# save a result.xlsx sheet


from openpyxl import Workbook
from openpyxl import load_workbook

#---------- create new sheet in workbook
handler1 = Workbook()
sheet1 = handler1.active
#--------------------------------------


#----------Load a workbook and make sheet active ----------
handler2 = load_workbook(filename = 'test.xlsx')
sheet2 = handler2.active
#----------------------------------------------------------

sheet1.cell(row=1, column=1).value = "NAME"
sheet1.cell(row=1, column=3).value = "RESULT" 
i=2
while sheet2.cell(row=i, column=3).value:

    sheet1.cell(row=i, column=1).value = sheet2.cell(row=i, column=1).value
    #  sheet1.cell(row=i, column=2).value = sheet2.cell(row=i, column=2).value
    if sheet2.cell(row=i, column=3).value >= 250:
       sheet1.cell(row=i, column=3).value = "PASS"
    else:
       sheet1.cell(row=i, column=3).value = "FAIL"        
    i= i + 1



handler1.save("test2.xlsx")
